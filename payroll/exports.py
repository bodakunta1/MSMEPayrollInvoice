from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def build_payroll_wage_register_workbook(payroll_run):
    """
    Builds an Excel wage register for one PayrollRun.

    Source:
    PayrollRun -> PayrollLine records

    This function does not recalculate payroll.
    It only exports already calculated payroll data.
    """

    wb = Workbook()
    ws = wb.active
    ws: Worksheet = wb.active
    ws.title = "Wage Register"

    company = payroll_run.company
    po = payroll_run.po
    cycle = payroll_run.payroll_cycle

    # Styles
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=11)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    section_fill = PatternFill("solid", fgColor="E5E7EB")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    money_format = '₹ #,##0.00'
    number_format = '0.00'

    # Title area
    ws.merge_cells("A1:U1")
    ws["A1"] = company.name
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:U2")
    ws["A2"] = "PO-wise Monthly Wage Register"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "PO Number"
    ws["B4"] = po.po_number

    ws["D4"] = "Payroll Cycle"
    ws["E4"] = cycle.name

    ws["G4"] = "Period"
    ws["H4"] = f"{cycle.period_start} to {cycle.period_end}"

    ws["J4"] = "Run Number"
    ws["K4"] = payroll_run.run_number

    ws["A5"] = "Location"
    ws["B5"] = po.location or ""

    ws["D5"] = "Department"
    ws["E5"] = po.department or ""

    ws["G5"] = "Status"
    ws["H5"] = payroll_run.status

    for cell in ["A4", "D4", "G4", "J4", "A5", "D5", "G5"]:
        ws[cell].font = subtitle_font
        ws[cell].fill = section_fill

    # Table headers
    headers = [
        "S.No",
        "Labour Code",
        "Labourer Name",
        "Skill Group",
        "Working Days",
        "Basic Hours",
        "OT Hours",
        "Basic Wage",
        "JAC Allowance",
        "Other Cash",
        "OT Amount",
        "Additional Allowance",
        "Gross Wage",
        "PF Deduction",
        "ESI Deduction",
        "Other Advance",
        "Festival Advance",
        "Total Deductions",
        "Net Pay",
        "Remarks",
        "Signature",
    ]

    header_row = 7

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    lines = payroll_run.lines.all().order_by("labourer_name")

    start_row = header_row + 1

    for index, line in enumerate(lines, start=1):
        row = start_row + index - 1

        values = [
            index,
            line.labour_code,
            line.labourer_name,
            line.skill_group_name,
            line.working_days,
            line.basic_hours,
            line.overtime_hours,
            line.basic_wage,
            line.jac_allowance,
            line.other_cash,
            line.overtime_amount,
            line.additional_allowance,
            line.gross_wage,
            line.pf_deduction,
            line.esi_deduction,
            line.other_advance,
            line.festival_advance,
            line.total_deductions,
            line.net_pay,
            line.calculation_notes,
            "",
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_num in [5, 6, 7]:
                cell.number_format = number_format

            if col_num in [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Totals row
    total_row = start_row + lines.count()

    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = subtitle_font
    ws.cell(row=total_row, column=1).fill = section_fill

    # Merge TOTAL label from A to G
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)

    total_columns = {
        8: payroll_run.total_basic_wage,
        9: payroll_run.total_jac_allowance,
        10: payroll_run.total_other_cash,
        11: payroll_run.total_overtime,
        12: payroll_run.total_additional_allowance,
        13: payroll_run.total_gross_wage,
        14: payroll_run.total_pf_deduction,
        15: payroll_run.total_esi_deduction,
        16: payroll_run.total_other_advance,
        17: payroll_run.total_festival_advance,
        18: payroll_run.total_deductions,
        19: payroll_run.total_net_pay,
    }

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_num)
        cell.border = thin_border
        cell.fill = section_fill
        cell.font = subtitle_font

        if col_num in total_columns:
            cell.value = total_columns[col_num]
            cell.number_format = money_format
            cell.alignment = Alignment(horizontal="right")

    # Freeze headers and add filter
    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:U{total_row}"

    # Column widths
    widths = {
        "A": 8,
        "B": 14,
        "C": 28,
        "D": 18,
        "E": 14,
        "F": 14,
        "G": 12,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 20,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 15,
        "Q": 18,
        "R": 18,
        "S": 15,
        "T": 30,
        "U": 18,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Page setup for printing
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.oddHeader.center.text = company.name
    ws.oddHeader.right.text = "Wage Register"
    ws.oddFooter.center.text = "Page &P of &N"

    return wb